"""Pro Excel formatter for DSH wide-research matrices.

Runs inside a Daytona sandbox. Reads a JSON matrix from ``--input`` and writes
a formatted ``.xlsx`` workbook to ``--output``. By default PII fields such as
``content``, ``email``, ``phone``, and ``tax_id`` are stripped from sources.
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
from datetime import UTC, datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

MAX_FILE_SIZE_BYTES = 10 * 1024 * 1024
PII_KEYS = frozenset({"content", "email", "phone", "tax_id", "address"})
SOURCE_WHITELIST = frozenset({"title", "url", "source_type"})


def _redact_source(source: Any, include_pii: bool) -> dict[str, Any]:
    """Return a source row limited to safe fields."""
    if not isinstance(source, dict):
        return {"title": "", "url": "", "source_type": "web"}
    if include_pii:
        return {k: v for k, v in source.items() if v is not None}
    return {
        k: v for k, v in source.items() if k in SOURCE_WHITELIST and v is not None
    }


def _mask_pii_in_topics(topics: list[Any]) -> list[str]:
    """Best-effort masking for PII tokens that might leak into topic text."""
    masked: list[str] = []
    for topic in topics:
        if not isinstance(topic, str):
            topic = str(topic)
        # Mask email-like and phone-like tokens with word boundaries.
        topic = re.sub(r"\b[\w.-]+@[\w.-]+\.\w+\b", "[EMAIL]", topic)
        topic = re.sub(r"\b\d{9,11}\b", "[PHONE]", topic)
        masked.append(topic)
    return masked


def _validate_input(data: Any) -> dict[str, Any]:
    if not isinstance(data, dict):
        raise ValueError("Input must be a JSON object")
    for key in ("topics", "sources", "matrix"):
        if key not in data:
            raise ValueError(f"Input is missing required key: {key}")
    return data


TRUTHY_STRINGS = frozenset({"1", "true", "yes", "y", "t"})
FALSY_STRINGS = frozenset({"", "false", "0", "no", "n", "f"})


def _bool_from_raw(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value != 0
    if isinstance(value, str):
        cleaned = value.strip().lower()
        if cleaned in TRUTHY_STRINGS:
            return True
        if cleaned in FALSY_STRINGS:
            return False
        # Unknown strings should not silently become True.
        return False
    return bool(value)


def _write_workbook(input_path: str, output_path: str, include_pii: bool) -> None:
    with open(input_path, encoding="utf-8") as f:
        data = _validate_input(json.load(f))

    topics = [str(t) for t in data.get("topics", []) if t is not None]
    sources = [_redact_source(s, include_pii) for s in data.get("sources", [])]
    raw_matrix = data.get("matrix", [])

    # Build a clean boolean matrix aligned with (sources x topics).
    matrix: list[list[bool]] = []
    for row in raw_matrix:
        if not isinstance(row, list):
            row = []
        clean_row = [_bool_from_raw(v) for v in row[: len(topics)]]
        while len(clean_row) < len(topics):
            clean_row.append(False)
        matrix.append(clean_row)
    while len(matrix) < len(sources):
        matrix.append([False] * len(topics))

    wb = Workbook()
    # Remove default sheet.
    wb.remove(wb.active)  # type: ignore[arg-type]

    # --- Summary tab ---
    summary = wb.create_sheet("Summary")
    summary.append(["Wide Research Deliverable"])
    summary.append(["Generated at", datetime.now(UTC).isoformat()])
    summary.append(["Topics", len(topics)])
    summary.append(["Sources", len(sources)])
    summary.append(["PII included", "Yes" if include_pii else "No"])
    summary["A1"].font = Font(bold=True, size=14)

    # --- Sources tab ---
    sources_sheet = wb.create_sheet("Sources")
    headers = ["#", "Title", "URL", "Source Type"] + (
        ["Content"] if include_pii and any(s.get("content") for s in sources) else []
    )
    sources_sheet.append(headers)
    for idx, src in enumerate(sources, start=1):
        row = [idx, src.get("title", ""), src.get("url", ""), src.get("source_type", "")]
        if include_pii:
            row.append(src.get("content", ""))
        sources_sheet.append(row)
    _apply_header_style(sources_sheet)
    _apply_autofilter(sources_sheet)
    _auto_width(sources_sheet)

    # --- Topics tab ---
    topics_sheet = wb.create_sheet("Topics")
    topics_sheet.append(["#", "Topic", "Source Count"])
    for idx, topic in enumerate(_mask_pii_in_topics(topics), start=1):
        col = get_column_letter(idx + 2)  # C, D, E...
        formula = f"=COUNTIF(Matrix!{col}:{col}, TRUE)"
        topics_sheet.append([idx, topic, formula])
    _apply_header_style(topics_sheet)
    _apply_autofilter(topics_sheet)
    _auto_width(topics_sheet)

    # --- Matrix tab ---
    matrix_sheet = wb.create_sheet("Matrix")
    matrix_sheet.append(["Source \\ Topic", *topics])
    for src, row in zip(sources, matrix, strict=False):
        matrix_sheet.append([src.get("title", ""), *row])
    _apply_header_style(matrix_sheet)
    _apply_autofilter(matrix_sheet)
    _auto_width(matrix_sheet)
    _apply_boolean_fill(matrix_sheet)

    wb.save(output_path)

    file_size = os.path.getsize(output_path)
    if file_size > MAX_FILE_SIZE_BYTES:
        os.remove(output_path)
        raise RuntimeError(f"Output file too large: {file_size} bytes (max {MAX_FILE_SIZE_BYTES})")

    print(f"Wrote {output_path} ({file_size} bytes)")


def _apply_header_style(sheet: Any) -> None:
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9E1F2")


def _apply_autofilter(sheet: Any) -> None:
    if sheet.max_row >= 1 and sheet.max_column >= 1:
        sheet.auto_filter.ref = sheet.dimensions


def _auto_width(sheet: Any) -> None:
    for column in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                value = str(cell.value) if cell.value is not None else ""
                max_length = max(max_length, len(value))
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 80)
        sheet.column_dimensions[column_letter].width = adjusted_width


def _apply_boolean_fill(sheet: Any) -> None:
    green = PatternFill("solid", fgColor="C6EFCE")
    for row in sheet.iter_rows(min_row=2, min_col=2):
        for cell in row:
            if cell.value is True:
                cell.fill = green


def main() -> int:
    parser = argparse.ArgumentParser(description="Format a wide-research matrix into .xlsx")
    parser.add_argument("--input", required=True, help="Path to input JSON matrix")
    parser.add_argument("--output", required=True, help="Path to output .xlsx file")
    parser.add_argument("--include-pii", action="store_true", help="Include PII fields")
    args = parser.parse_args()

    try:
        _write_workbook(args.input, args.output, args.include_pii)
        return 0
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    sys.exit(main())
