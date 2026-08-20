"""Unit tests for the DSH deliver subgraph and Pro Excel formatter template."""

from __future__ import annotations

import json
import subprocess
import sys
from pathlib import Path
from unittest.mock import AsyncMock, MagicMock, patch

import pytest

from app.tasks.dsh_worker_deliver_subgraph import (
    DshDeliverSubgraph,
    _redact_matrix,
)

SAMPLE_MATRIX = {
    "topics": ["framework", "model"],
    "sources": [
        {
            "title": "LangChain",
            "url": "https://langchain.com",
            "source_type": "web",
            "phone": "0901234567",
            "email": "test@example.com",
        },
        {
            "title": "LangGraph",
            "url": "https://langgraph.com",
            "source_type": "web",
        },
    ],
    "matrix": [
        [True, False],
        [False, True],
    ],
}


def test_redact_matrix_strips_pii() -> None:
    safe = _redact_matrix(SAMPLE_MATRIX)
    assert safe["topics"] == SAMPLE_MATRIX["topics"]
    assert safe["matrix"] == SAMPLE_MATRIX["matrix"]
    assert len(safe["sources"]) == 2
    for src in safe["sources"]:
        assert "phone" not in src
        assert "email" not in src
        assert "title" in src
        assert "url" in src
        assert "source_type" in src


def test_formatter_template_creates_xlsx(tmp_path: Path) -> None:
    script = Path("scripts") / "sandbox_pro_excel_template.py"
    if not script.exists():
        pytest.skip("formatter template not found")

    input_path = tmp_path / "matrix.json"
    output_path = tmp_path / "out.xlsx"
    input_path.write_text(json.dumps(SAMPLE_MATRIX))

    result = subprocess.run(
        [sys.executable, str(script), "--input", str(input_path), "--output", str(output_path)],
        capture_output=True,
        text=True,
        check=False,
    )
    assert result.returncode == 0, result.stderr
    assert output_path.exists()
    assert output_path.stat().st_size < 10 * 1024 * 1024

    # Light structural check with openpyxl.
    from openpyxl import load_workbook

    wb = load_workbook(output_path)
    assert {"Summary", "Sources", "Topics", "Matrix"}.issubset(set(wb.sheetnames))

    sources_sheet = wb["Sources"]
    # The PII-free source sheet should not contain the phone number.
    for row in sources_sheet.iter_rows(values_only=True):
        for cell in row:
            if isinstance(cell, str):
                assert "0901234567" not in cell
                assert "test@example.com" not in cell


@pytest.mark.asyncio
async def test_deliver_subgraph_skips_without_matrix() -> None:
    result = await DshDeliverSubgraph().run("mission-1", {})
    assert result is None


@pytest.mark.asyncio
async def test_deliver_subgraph_runs_sandbox_and_returns_deliverable(tmp_path: Path) -> None:
    fake_local = tmp_path / "wide_research_output.xlsx"
    fake_local.write_bytes(b"fake-xlsx")

    fake_sandbox = MagicMock()
    fake_sandbox.aexecute = AsyncMock(
        return_value=MagicMock(output="Wrote /documents/wide_research_output.xlsx", exit_code=0)
    )

    with patch(
        "app.tasks.dsh_worker_deliver_subgraph.get_or_create_sandbox",
        return_value=(fake_sandbox, True),
    ) as mock_get, patch(
        "app.tasks.dsh_worker_deliver_subgraph.sync_files_to_sandbox",
        new_callable=AsyncMock,
    ) as mock_sync, patch(
        "app.tasks.dsh_worker_deliver_subgraph.persist_and_delete_sandbox",
        return_value=None,
    ) as mock_persist, patch(
        "app.tasks.dsh_worker_deliver_subgraph.get_local_sandbox_file",
        return_value=b"fake-xlsx",
    ) as mock_get_local:
        result = await DshDeliverSubgraph().run("mission-1", {"wide_research_matrix": SAMPLE_MATRIX})

    assert result is not None
    assert result["type"] == "xlsx"
    assert result["filename"] == "wide_research_output.xlsx"
    assert result["include_pii"] is False
    assert result["size"] == 9
    assert result["sandbox_path"] == "/documents/wide_research_output.xlsx"
    mock_get.assert_called_once_with("mission-1")
    mock_sync.assert_awaited_once_with("mission-1", mock_sync.call_args[0][1], fake_sandbox, True)
    mock_persist.assert_called_once_with("mission-1", ["/documents/wide_research_output.xlsx"])
    mock_get_local.assert_called_once_with("mission-1", "/documents/wide_research_output.xlsx")
    fake_sandbox.aexecute.assert_awaited_once()
